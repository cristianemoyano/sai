from decimal import Decimal
from django.shortcuts import render
from search_views.search import SearchListView
from django.db import transaction
from django.views.generic import ListView, DetailView
from django.core import serializers
from django.http import HttpResponse
from .resources import ProductResource
from .forms import (
    ProductCreateForm,
    ProductEditForm,
    ProductSearchForm,
    ProductFilter,
    ProductCategoryForm,
    ProductStoreForm,
    PriceForm,
)
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy

from .models import (
    Product,
    ProductStatus,
    ProductUnit,
    ProductCategory,
    ProductStore,
    Price,
    Currency,
)
from purchases.models import (
    Provider,
    Purchase,
)
from sales.models import (
    Customer,
    Order,
)
from rest_framework import viewsets
from .serializers import ProductSerializer
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required

from django.db.models import Sum
from main.auth import GroupRequiredMixin
from main.utils import get_general_stock_status, get_historical_sales_amount_by_month
from openpyxl import load_workbook


@login_required
def dashboard(request):
    product_alerts = get_general_stock_status()
    products = Product.objects.all().count()
    providers = Provider.objects.all().count()
    customers = Customer.objects.all().count()
    purchases = Purchase.objects.all().count()
    orders = Order.objects.all().count()
    orders_sum = Order.objects.aggregate(Sum('paid_amount'))
    purchases_sum = Purchase.objects.aggregate(Sum('paid_amount'))

    purchases_amount = purchases_sum.get('paid_amount__sum')
    orders_amount = orders_sum.get('paid_amount__sum')

    if not purchases_amount:
        purchases_amount = 0
    if not orders_amount:
        orders_amount = 0
    
    historical_sales = get_historical_sales_amount_by_month()
    context = {
        'products': products,
        'providers': providers,
        'customers': customers,
        'purchases': purchases,
        'orders': orders,
        'orders_amount': orders_amount,
        'purchases_amount': purchases_amount,
        'products_in_danger': product_alerts.get('danger'),
        'products_in_alert': product_alerts.get('alert'),
        'products_in_info': product_alerts.get('info'),
        'products_in_success': product_alerts.get('success'),
        'historical_sales': historical_sales,
    }

    return render(request, 'product/dashboard.html', context)

# Product CRUD


class ProductList(LoginRequiredMixin, SearchListView):
    model = Product
    template_name = "product/product_list.html"

    # additional configuration for SearchListView
    form_class = ProductSearchForm
    filter_class = ProductFilter


class ProductView(LoginRequiredMixin, DetailView):
    model = Product


class ProductCreate(LoginRequiredMixin, CreateView):
    model = Product
    success_url = reverse_lazy('product_list')
    form_class = ProductCreateForm

    def form_valid(self, form):
        form.instance.user_id = self.request.user.id
        return super().form_valid(form)


class ProductUpdate(GroupRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Product
    success_url = reverse_lazy('product_list')
    form_class = ProductEditForm

    def form_valid(self, form):
        form.instance.user_id = self.request.user.id
        return super().form_valid(form)


class ProductDelete(GroupRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Product
    success_url = reverse_lazy('product_list')


def product_import_file(request):
    success_url = 'product/product_import.html'
    context = {}
    if request.method == 'POST':
        request_file = request.FILES["file"]
        if request_file.name.endswith('.csv'):
            context = process_csv(request_file, request.user.id)
            return render(request, success_url, context)
        elif request_file.name.endswith(('xlsx', 'xls')):
            context = process_xlsx(request_file, request.user)
            return render(request, success_url, context)
        else:
            context['error'] = 'Formato no permitido. Sólo se permiten .xlsx, .csv'
            return render(request, success_url, context)
    return render(request, success_url, context)

def process_csv(request_file, user_id):
    product_fields = [
        'id',
        'created',
        'modified',
        'user',
        'name',
        'reference_code',
        'import_code',
        'description',
        'stock',
        'min_amount',
        'product_store',
        'product_category',
        'product_type',
        'product_status',
        'image_url'
    ]
    context = {}
    try:
        # if file is too large, return
        if request_file.multiple_chunks():
            context['error'] = 'Archivo demasiado grande.'
            return context
        file_data = request_file.read().decode("utf-8")
        lines = file_data.split("\r\n")
        for line in lines:
            fields = line.split(",")
            data_dict = {}
            i = 0
            for field in product_fields:
                if len(product_fields) == len(fields):
                    data_dict[field] = fields[i]
                    i += 1
            try:
                form = ProductCreateForm(data_dict)
                form.instance.user_id = user_id
                if form.is_valid():
                    form.save()
            except Exception as e:
                context['error'] = 'Error inesperado: {e}'.format(e=e)
                return context
        context['message'] = 'Productos creados correctamente.'
        return context
    except Exception as e:
        context['error'] = 'Error inesperado: {e}'.format(e=e)
        return context

def process_xlsx(request_file, user):
    COL_NAMES = {
        1:'reference_code',
        2:'name',
        3:'cost_price',
        4:'list_price',
        5:'price_a',
        6:'stock',
        7:'min_amount',
        8:'product_category',
    }
    START_ROW = 2
    START_COL = 1
    context = {}
    wb2 = load_workbook(request_file.file)
    sheet_obj = wb2.active
    max_row = sheet_obj.max_row + 1
    max_column = sheet_obj.max_column + 1
    data_sheet = []
    for row in range(START_ROW, max_row):
        data_row = {}
        for col in range(START_COL, max_column):
            cell_obj = sheet_obj.cell(row = row, column = col)
            data_row[COL_NAMES.get(col)] = cell_obj.value
        data_sheet.append(data_row)
    save_data_from_sheet(data_sheet, user)
    context['message'] = 'Productos creados correctamente.'
    return context

def save_data_from_sheet(data_sheet, user):
    for row in data_sheet:
        save_product(row, user)

@transaction.atomic
def save_product(row, user):
    try:
        product_category = ProductCategory.objects.get(name=row.get('product_category'))
    except Exception:
        product_category_data = {
            'name': row.get('product_category'),
            'code': row.get('product_category').upper(),
            'description': row.get('product_category'),
        }
        product_category = ProductCategory(**product_category_data)
        product_category.save()
    sid = transaction.savepoint()
    try:
        
        product_data = {
                'user': user,
                'name': row.get('name'),
                'reference_code': row.get('reference_code').upper(),
                'stock': Decimal(row.get('stock')),
                'product_unit': ProductUnit.objects.get(code='PIECE'),
                'min_amount': Decimal(row.get('min_amount')),
                'product_category': product_category,
                'product_status': ProductStatus.objects.get(code='available'),
                'image_url': '/static/images/no-image.png'
        }
        product = Product(**product_data)
        product.save()
        price_data = {
            'user': user,
            'product': product,
            'currency': Currency.objects.get(code='ARS'),
            'cost_price': Decimal(row.get('cost_price').split('$')[1]),
            'list_price': Decimal(row.get('list_price').split('$')[1]),
            'price_a': Decimal(row.get('price_a').split('$')[1]),
        }
        price = Price(**price_data)
        price.save()
    except Exception as e:
        print(e)
        try:
            transaction.savepoint_rollback(sid)
        except:
            pass

def product_export_csv(request):
    product_resource = ProductResource()
    dataset = product_resource.export()
    response = HttpResponse(dataset.csv, content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="producto.csv"'
    return response


def product_export_excel(request):
    product_resource = ProductResource()
    dataset = product_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="producto.xls"'
    return response

# END Product CRUD

# ProductCategory CRUD


class ProductCategoryList(LoginRequiredMixin, ListView):
    model = ProductCategory


class ProductCategoryView(DetailView):
    model = ProductCategory


class ProductCategoryCreate(LoginRequiredMixin, CreateView):
    model = ProductCategory
    success_url = reverse_lazy('product_category_list')
    form_class = ProductCategoryForm


class ProductCategoryUpdate(GroupRequiredMixin, LoginRequiredMixin, UpdateView):
    model = ProductCategory
    success_url = reverse_lazy('product_category_list')
    form_class = ProductCategoryForm

    def form_valid(self, form):
        form.instance.user_id = self.request.user.id
        return super().form_valid(form)


class ProductCategoryDelete(GroupRequiredMixin, LoginRequiredMixin, DeleteView):
    model = ProductCategory
    success_url = reverse_lazy('product_category_list')

# END ProductCategory CRUD

# ProductStore CRUD


class ProductStoreList(LoginRequiredMixin, ListView):
    model = ProductStore


class ProductStoreView(DetailView):
    model = ProductStore


class ProductStoreCreate(LoginRequiredMixin, CreateView):
    model = ProductStore
    success_url = reverse_lazy('product_store_list')
    form_class = ProductStoreForm


class ProductStoreUpdate(GroupRequiredMixin, LoginRequiredMixin, UpdateView):
    model = ProductStore
    success_url = reverse_lazy('product_store_list')
    form_class = ProductStoreForm

    def form_valid(self, form):
        form.instance.user_id = self.request.user.id
        return super().form_valid(form)


class ProductStoreDelete(GroupRequiredMixin, LoginRequiredMixin, DeleteView):
    model = ProductStore
    success_url = reverse_lazy('product_store_list')

# END ProductStore CRUD

# Price CRUD


class PriceList(LoginRequiredMixin, ListView):
    model = Price


class PriceView(DetailView):
    model = Price


class PriceCreate(LoginRequiredMixin, CreateView):
    model = Price
    success_url = reverse_lazy('product_price_list')
    form_class = PriceForm

    def form_valid(self, form):
        form.instance.user_id = self.request.user.id
        return super().form_valid(form)


class PriceUpdate(GroupRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Price
    success_url = reverse_lazy('product_price_list')
    form_class = PriceForm

    def form_valid(self, form):
        form.instance.user_id = self.request.user.id
        return super().form_valid(form)


class PriceDelete(GroupRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Price
    success_url = reverse_lazy('product_price_list')

# END Price CRUD


# API
class ProductAPIView(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Product.objects.all().order_by('-created')
    serializer_class = ProductSerializer
    lookup_url_kwarg = 'reference_code'

    def get_queryset(self):
        code = self.request.query_params.get(self.lookup_url_kwarg)
        products = Product.objects.all().order_by('-created')
        if code:
            products = Product.objects.filter(reference_code=code)
        return products
